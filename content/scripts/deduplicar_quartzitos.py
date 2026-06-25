import os
from openpyxl import load_workbook

def processar_catalogo_nuovolam():
    # --- 1. CONFIGURAÇÃO DE CAMINHOS ---
    pasta_trabalho = r'C:\Iven Stone\catalogo\quartzitos'
    nome_arquivo = 'relacao_quartizitos_v3.11.xlsx'
    caminho_completo = os.path.join(pasta_trabalho, nome_arquivo)

    if not os.path.exists(caminho_completo):
        print(f"❌ Erro: Arquivo não encontrado em:\n{caminho_completo}")
        return

    print(f"📂 Abrindo arquivo: {nome_arquivo}")

    # --- 2. ABERTURA DO WORKBOOK ---
    workbook = load_workbook(caminho_completo, data_only=True)
    abas = workbook.sheetnames

    print(f"📑 Abas encontradas no workbook: {abas}")

    # --- 3. VALIDAÇÃO E ACESSO DAS ABAS POR NOME ---
    nome_aba_origem = 'strings'
    nome_aba_destino = 'script_1'

    if nome_aba_origem not in abas:
        print(f"❌ Erro: A aba '{nome_aba_origem}' não foi encontrada.")
        return

    if nome_aba_destino not in abas:
        print(f"❌ Erro: A aba '{nome_aba_destino}' não foi encontrada.")
        return

    worksheet_origem = workbook[nome_aba_origem]
    worksheet_destino = workbook[nome_aba_destino]

    print(f"🔍 Aba de origem: '{nome_aba_origem}'")
    print(f"🎯 Aba de destino: '{nome_aba_destino}'")

    # --- 4. LEITURA E DEDUPLICAÇÃO ---
    # openpyxl usa índice base 1. Coluna D = 4, Coluna E = 5
    # Começamos na linha 2 para pular o cabeçalho
    combinacoes = []
    total_linhas = worksheet_origem.max_row

    for linha_num in range(2, total_linhas + 1):
        col_d = worksheet_origem.cell(row=linha_num, column=4).value
        col_e = worksheet_origem.cell(row=linha_num, column=5).value

        # Converte para string e trata valores None
        col_d_str = str(col_d).strip() if col_d is not None else ""
        col_e_str = str(col_e).strip() if col_e is not None else ""

        # Só considera linhas que tenham ao menos um valor preenchido
        if col_d_str or col_e_str:
            combinacoes.append((col_d_str, col_e_str))

    # Remove duplicatas preservando a ordem original de aparição
    itens_unicos = list(dict.fromkeys(combinacoes))

    print(f"\n📊 RESULTADO DA CURADORIA:")
    print(f"   • Total de linhas lidas (excluindo cabeçalho): {len(combinacoes)}")
    print(f"   • Total de combinações únicas (Produto + Acabamento): {len(itens_unicos)}")
    print(f"   • Duplicatas removidas: {len(combinacoes) - len(itens_unicos)}")

    # --- 5. ESCRITA NA ABA "script_1" ---
    # Limpa o conteúdo anterior da aba de destino
    for row in worksheet_destino.iter_rows():
        for cell in row:
            cell.value = None

    # Escreve o cabeçalho na linha 1
    worksheet_destino.cell(row=1, column=1, value="Nome do Produto (Coluna D)")
    worksheet_destino.cell(row=1, column=2, value="Acabamento (Coluna E)")

    # Escreve os dados únicos a partir da linha 2
    for idx, (produto, acabamento) in enumerate(itens_unicos, start=2):
        worksheet_destino.cell(row=idx, column=1, value=produto)
        worksheet_destino.cell(row=idx, column=2, value=acabamento)

    # Ajusta automaticamente a largura das colunas para melhor visualização
    worksheet_destino.column_dimensions['A'].width = 40
    worksheet_destino.column_dimensions['B'].width = 30

    # --- 6. SALVAMENTO ---
    workbook.save(caminho_completo)
    print(f"\n✅ Sucesso! Arquivo salvo com os dados únicos na aba '{nome_aba_destino}'.")

if __name__ == '__main__':
    processar_catalogo_nuovolam()